#!/usr/bin/env python3
"""Generate demo seed data from Connect Dots masterlist + scripts/seed-config.json.

Lean scope: 6 projects, ~2–3 real style codes each — enough to demo the flow.
Re-run after xlsx or config changes:  npm run import-masterlist
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Install openpyxl: pip3 install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "scripts" / "seed-config.json"
OUT_PROJECTS = ROOT / "src/lib/mock/generated/demo-projects.ts"
OUT_JOBS = ROOT / "src/lib/mock/generated/demo-jobs.ts"
OUT_STYLES = ROOT / "src/lib/reference/styles-seed.ts"

SHEET_TO_JOB_CATEGORY = {
    "TEES": "Tee",
    "SWEATSHIRTS": "Sweatshirt",
    "PANTS": "Pants",
    "SHORTS": "Shorts",
    "JACKETS": "Jacket",
    "KNITS": "Knit",
    "SKIRTS": "Skirt",
    "HEADWEAR": "Hat",
    "DRESS": "Dress",
    "UNDERWEAR": "Underwear",
    "BAGS": "Bag",
    "ACCESSORIES": "Accessory",
    "Home Goods": "Accessory",
    "Paper Goods": "Accessory",
    "SKATEBOARD": "Accessory",
    "STICKERS": "Accessory",
}

SHEET_TO_PROJECT_CATEGORY = {
    "TEES": "TEE",
    "SWEATSHIRTS": "SWEATSHIRT",
    "PANTS": "PANTS",
    "SHORTS": "SHORTS",
    "JACKETS": "JACKET",
    "KNITS": "KNIT",
    "SKIRTS": "SKIRT",
    "HEADWEAR": "HEADWEAR",
    "DRESS": "DRESS",
    "UNDERWEAR": "UNDERWEAR",
    "BAGS": "BAG",
    "ACCESSORIES": "ACCESSORIES",
    "Home Goods": "HOME GOODS",
    "Paper Goods": "PAPER GOODS",
    "SKATEBOARD": "SKATEBOARD",
    "STICKERS": "STICKERS",
}

JOB_TYPE_BY_CATEGORY = {
    "Bag": "branding",
    "Accessory": "branding",
    "Hat": "branding",
    "Sticker": "branding",
}


def load_json(path: Path) -> dict:
    return json.loads(path.read_text())


def resolve_xlsx(config: dict) -> Path:
    raw = config.get("xlsxPath", "")
    p = (ROOT / raw).resolve() if not Path(raw).is_absolute() else Path(raw)
    if not p.exists():
        fallback = Path.home() / "Downloads" / "Connect Dots - Style, Color & Client Code Masterlist.xlsx"
        if fallback.exists():
            return fallback
        raise FileNotFoundError(f"Masterlist not found: {p}")
    return p


def load_clients(wb) -> dict[str, str]:
    ws = wb["Client Codes"]
    out: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        code, name = row[0], row[1]
        if code and name and str(code) != "CLIENT CODES":
            out[str(code).strip().upper()] = str(name).strip()
    return out


def load_style_index(wb) -> dict[str, dict]:
    index: dict[str, dict] = {}
    for sheet in wb.sheetnames:
        if "Style Codes" not in sheet or sheet == "Style Code Key":
            continue
        cat_sheet = sheet.replace(" - Style Codes", "")
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:
                continue
            code = str(row[0] or "").strip().upper()
            if not code:
                continue
            index[code] = {
                "styleCode": code,
                "clientLabel": str(row[1] or "").strip(),
                "styleName": str(row[2] or "").strip() or code,
                "note": str(row[3] or "").strip() if len(row) > 3 and row[3] else "",
                "categorySheet": cat_sheet,
            }
    return index


def iso(date_str: str) -> str:
    return f"{date_str}T12:00:00.000Z"


def po_number(code: str, year: int, seq: int, month: int = 5) -> str:
    return f"{code}-{year}-{month:02d}-{seq:03d}"


def ts_literal(s: str) -> str:
    return json.dumps(s)


def null_project_tail() -> str:
    return """
    client_meeting_date: null,
    client_assets_received_date: null,
    cs_tech_pack_request_date: null,
    cs_tech_pack_due_date: null,
    cs_tech_pack_assigned_member: null,
    cs_tech_pack_complete_date: null,
    artwork_tech_pack_request_date: null,
    artwork_tech_pack_due_date: null,
    artwork_tech_pack_assigned_member: null,
    artwork_tech_pack_complete_date: null,
    artwork_design_client_approval_date: null,
    dev_prod_assigned_team_member: null,
    tp_sent_date: null,
    references_sent_date: null,
    quote_requested_date: null,
    vendor_costing_received_date: null,
    cost_sheet_prepared_date: null,
    estimate_sent_date: null,
    costing_approved: null,
    dye_vendor: null,
    lab_dip_request_date: null,
    lab_dip_due_date: null,
    lab_dip_received_date: null,
    lab_dip_approval_status: null,
    print_embroidery_vendor: null,
    strike_off_request_date: null,
    strike_off_due_date: null,
    strike_off_received_date: null,
    strike_off_approval_status: null,
    bulk_fabric_approval_date: null,
    bulk_trim_approval_date: null,
    new_product_request_date: null,
    barcodes_sent_to_vendor_date: null,
    top_due_date: null,
    top_approved_date: null,
    bulk_target_delivery_date: null,
    ex_factory_date: null,
    shipping_terms: null,
    shipping_method: null,
    packing_list_received_date: null,
    tracking_bol_number: null,
    packing_list_sent_to_client_date: null,
    client_received_date: null,"""


def main() -> None:
    config = load_json(CONFIG_PATH)
    xlsx = resolve_xlsx(config)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    clients = load_clients(wb)
    styles = load_style_index(wb)
    year = int(config.get("year", datetime.now().year))
    projects_cfg = config["projects"]

    missing: list[str] = []
    for p in projects_cfg:
        for j in p["jobs"]:
            code = j["styleCode"].upper()
            if code not in styles:
                missing.append(code)
    if missing:
        wb.close()
        raise SystemExit(f"Style codes not found in masterlist: {', '.join(missing)}")

    project_blocks: list[str] = []
    job_blocks: list[str] = []
    style_entries: list[str] = []
    seen_styles: set[str] = set()

    for idx, pcfg in enumerate(projects_cfg, start=1):
        code = pcfg["clientCode"].upper()
        client_name = clients.get(code)
        if not client_name:
            wb.close()
            raise SystemExit(f"Client code not in masterlist: {code}")

        first_job = pcfg["jobs"][0]
        first_style = styles[first_job["styleCode"].upper()]
        cat_sheet = first_style["categorySheet"]
        job_cat = SHEET_TO_JOB_CATEGORY.get(cat_sheet, "Tee")
        proj_cat = SHEET_TO_PROJECT_CATEGORY.get(cat_sheet, "TEE")

        handoff = iso("2026-04-08") if idx == 1 else iso(f"2026-0{min(idx, 9)}-15")
        due = iso("2026-07-15") if idx == 1 else iso(f"2026-0{min(idx + 1, 9)}-20")
        status_date = iso("2026-05-10") if pcfg["status"] == "IN-PROGRESS" else None

        project_blocks.append(
            f"""  {{
    id: {idx},
    name: {ts_literal(client_name)},
    description: {ts_literal(f"Connect Dots program — {first_style['styleName']}")},
    project_number: {ts_literal(first_style["styleCode"])},
    po_number: {ts_literal(po_number(code, year, 1))},
    project_hand_off_date: {ts_literal(handoff)},
    due_date: {ts_literal(due)},
    client: {{ id: {idx}, name: {ts_literal(client_name)}, avatar_url: null }},
    status: {ts_literal(pcfg["status"])},
    status_overview: {ts_literal(pcfg.get("statusOverview"))},
    status_update_date: {json.dumps(status_date)},
    style_number: {ts_literal(first_style["styleCode"])},
    style_name: {ts_literal(first_style["styleName"])},
    category: {ts_literal(proj_cat)},
    type: {ts_literal("PRINT / DECORATION ON BLANKS")},
    lead_vendor: {ts_literal(pcfg.get("leadVendor"))},
    colorways: [],
    in_development: [],
    lead_team_member: {ts_literal(pcfg.get("leadTeamMember"))},{null_project_tail()}
  }}"""
        )

        for j_idx, jcfg in enumerate(pcfg["jobs"]):
            sc = jcfg["styleCode"].upper()
            st = styles[sc]
            cat = SHEET_TO_JOB_CATEGORY.get(st["categorySheet"], "Tee")
            slug = re.sub(r"[^a-z0-9]+", "-", sc.lower()).strip("-")
            job_id = f"job-{idx}-{slug}"
            job_type = JOB_TYPE_BY_CATEGORY.get(cat, "print_production")
            type_label = "BRANDING" if job_type == "branding" else "PRINT PRODUCTION"

            job_blocks.append(
                f"""  {{
    id: {ts_literal(job_id)},
    project_id: {idx},
    name: {ts_literal(st["styleName"])},
    subtitle: {ts_literal(st["categorySheet"])},
    type: {ts_literal(type_label)},
    job_type: {ts_literal(job_type)},
    lead_vendor: {ts_literal(pcfg.get("leadVendor", "Millworks Collective"))},
    category: {ts_literal(cat)},
    style_number: {ts_literal(sc)},
    status: {ts_literal(jcfg["status"])},
    due_date: {ts_literal(due)},
    updated_at: {ts_literal(iso("2026-05-12"))},
    timeline: {jcfg.get("timeline", "lagging")}JobTimeline(),
    scope_kind: "original",
  }}"""
            )

            if sc not in seen_styles:
                seen_styles.add(sc)
                style_entries.append(
                    f'  {{ styleCode: {ts_literal(sc)}, clientCode: {ts_literal(code)}, clientName: {ts_literal(client_name)}, styleName: {ts_literal(st["styleName"])}, category: {ts_literal(cat)}, note: {ts_literal(st["note"]) or "null"} }},'
                )

    wb.close()

    OUT_PROJECTS.parent.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    projects_body = ",\n".join(project_blocks)
    jobs_body = ",\n".join(job_blocks)

    OUT_PROJECTS.write_text(
        f"""/** AUTO-GENERATED — npm run import-masterlist ({generated_at}) */
import type {{ Project }} from "@/lib/types/project";

export const demoProjects: Project[] = [
{projects_body},
];
"""
    )

    OUT_JOBS.write_text(
        f"""/** AUTO-GENERATED — npm run import-masterlist ({generated_at}) */
import type {{ ProjectJob }} from "@/lib/types/wip";

/** Timeline helpers injected by project-jobs.ts */
export type DemoJobTimelineFn = () => ProjectJob["timeline"];

export function buildDemoJobs(
  oliveJobTimeline: DemoJobTimelineFn,
  laggingJobTimeline: DemoJobTimelineFn,
): ProjectJob[] {{
  return [
{jobs_body},
  ];
}}
"""
    )

    OUT_STYLES.write_text(
        f"""/** AUTO-GENERATED demo styles only — full catalog deferred to backend. */
export type SeedStyleRow = {{
  styleCode: string;
  clientCode: string;
  clientName: string;
  styleName: string;
  category: string;
  note: string | null;
}};

/** Styles used in the 6 demo projects ({len(seen_styles)} rows). */
export const SEED_STYLES: SeedStyleRow[] = [
{chr(10).join(style_entries)}
];
"""
    )

    print(f"Wrote {OUT_PROJECTS.relative_to(ROOT)}")
    print(f"Wrote {OUT_JOBS.relative_to(ROOT)}")
    print(f"Wrote {OUT_STYLES.relative_to(ROOT)}")
    print(f"Projects: {len(projects_cfg)}, Jobs: {len(job_blocks)}, Styles: {len(seen_styles)}")


if __name__ == "__main__":
    main()
